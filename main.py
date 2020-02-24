import datetime
import os

import xlsxwriter
import openpyxl
from selenium.common.exceptions import WebDriverException

from core.abstract import Record
from sched.demorgen import demo
from util.util import schedule_headers_map, alpha, make_soup

from enrich import imdb as IMDB
from enrich import wiki as WIKI

today = datetime.date.today()


def records_to_schedule_excel_sheet(records: [Record], row_headers):
    year = datetime.datetime.now().year
    workbook_name = f"core/schedule_{year}.xlsx"

    if os.path.isfile(workbook_name):
        existing_workbook = openpyxl.load_workbook(workbook_name)
        ws = existing_workbook.active
        max_row = ws.max_row

        for row_number, record in enumerate(records):
            for c, v in enumerate(row_headers):
                cell_name = alpha[c] + str(row_number+max_row+1)
                ws[cell_name] = str(getattr(record, v))

        existing_workbook.save(workbook_name)
        existing_workbook.close()
    else:
        workbook = xlsxwriter.Workbook(workbook_name)  # Create new workbook for the year
        worksheet = workbook.add_worksheet()

        for c, h in enumerate(row_headers.keys()):
            worksheet.write(alpha[c] + str(1), row_headers[h])  # row_headers[h] to get the formatted row header

        for row, record in enumerate(records):
            for c, field_name in enumerate(row_headers.keys()):
                worksheet.write(alpha[c] + str(row + 2), str(getattr(record, field_name)))

        workbook.close()


def add_schedule_info_to_record(record: Record):      # Return if page was found & information added

    imdb_page_url = IMDB.get_page_url_from_title(record.title)
    parsed = False

    if imdb_page_url:
        page_soup = make_soup(imdb_page_url)
        IMDB.add_schedule_info_to_record(record, page_soup)
        parsed = True

    if not parsed:
        parsed = WIKI.add_en_info(record)

    if not parsed:
        WIKI.add_nl_info(record)


def get_schedule_records() -> [Record]:

    demoScraper = demo()
    records = demoScraper.get_records()

    for record in records:
        try:
            add_schedule_info_to_record(record)
            #print("Trying to add info for record with title:  " + record.title)
        except Exception as e:
            print("Error when adding info to record: " + str(e))
            continue
    return records


def run():
    try:
        records = get_schedule_records()
        records_to_schedule_excel_sheet(records, schedule_headers_map)
        print("\nWriting schedule to root directory... ")
    except WebDriverException:
        print("Selenium client not reachable, exiting...")


if __name__ == '__main__':
    run()






